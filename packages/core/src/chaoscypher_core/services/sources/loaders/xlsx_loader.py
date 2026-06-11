# Copyright (C) 2024-2026 Chaos Cypher, Inc.
# SPDX-License-Identifier: AGPL-3.0-only

"""Microsoft Excel (.xlsx / .xlsm) loader.

Walks every sheet, treats the sheet name as a section heading, and
flattens each non-empty row to a ``col1 | col2 | …`` line. Uses
``data_only=True`` so cached formula values come out as numbers/strings
rather than ``=SUM(A1:A3)``; ``read_only=True`` keeps memory bounded
for very large books.
"""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import TYPE_CHECKING, Any

import structlog
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from chaoscypher_core.exceptions import ValidationError
from chaoscypher_core.plugins import PluginMetadata


if TYPE_CHECKING:
    from chaoscypher_core.settings import EngineSettings


logger = structlog.get_logger(__name__)


class XLSXLoader:
    """Loader for Microsoft Excel ``.xlsx`` / ``.xlsm`` files."""

    @property
    def supported_extensions(self) -> list[str]:
        """File extensions this loader supports."""
        return [".xlsx", ".xlsm", ".XLSX", ".XLSM"]

    @property
    def metadata(self) -> PluginMetadata:
        """Plugin metadata."""
        return PluginMetadata(
            plugin_id="xlsx",
            name="Excel XLSX Loader",
            description="Loads Microsoft Excel .xlsx / .xlsm files.",
            version="1.0.0",
            author="ChaosCypher",
            category="loader",
            builtin=True,
            origin="builtin",
            tags=["document", "office", "spreadsheet"],
        )

    def __init__(self, settings: EngineSettings | None = None) -> None:
        """Initialize XLSX loader.

        Args:
            settings: Settings instance (not currently used).
        """
        self.settings = settings

    def load_document(self, filepath: str) -> list[dict[str, Any]]:
        """Load an Excel workbook.

        Args:
            filepath: Path to a .xlsx / .xlsm file.

        Returns:
            A single-item list with all sheets concatenated and metadata
            listing per-sheet row counts.

        Raises:
            ValidationError: If the file is not a valid XLSX/XLSM
                workbook (corrupt zip container, invalid OOXML, etc.).
        """
        path = Path(filepath)
        try:
            wb = load_workbook(str(path), data_only=True, read_only=True)
        except (InvalidFileException, zipfile.BadZipFile, KeyError) as exc:
            msg = f"File '{path.name}' is not a valid XLSX: {exc}"
            raise ValidationError(msg, field="content") from exc

        parts: list[str] = []
        sheets: list[dict[str, Any]] = []
        rows_skipped = 0
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"\n[Sheet: {sheet_name}]")
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    values = [str(cell) if cell is not None else "" for cell in row]
                    if not any(v.strip() for v in values):
                        rows_skipped += 1
                        continue
                    parts.append(" | ".join(values))
                    row_count += 1
                sheets.append({"name": sheet_name, "rows": row_count})
        finally:
            wb.close()

        content = "\n".join(parts)

        logger.info(
            "xlsx_loaded",
            filepath=str(path),
            sheet_count=len(sheets),
            row_total=sum(s["rows"] for s in sheets),
            rows_skipped=rows_skipped,
            character_count=len(content),
        )

        return [
            {
                "content": content,
                "metadata": {
                    "source": str(path),
                    "extraction_method": "xlsx",
                    "sheets": sheets,
                    "content_type": (
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    ),
                    "loader_xlsx_rows_skipped": rows_skipped,
                },
            }
        ]

    def supports_ocr(self) -> bool:
        """XLSX files don't need OCR."""
        return False


__all__ = ["XLSXLoader"]
